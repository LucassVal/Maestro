#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
from .extrator_base_SRC_GL import ExtratorBase

try:
    import openpyxl
except ImportError:
    openpyxl = None

class ExtratorXLSX(ExtratorBase):
    def __init__(self):
        super().__init__(nome="ExtratorXLSX")

    def executar(self, entrada):
        caminho = entrada.get("caminho")
        valido, msg = self.validar_arquivo(caminho)
        if not valido:
            return {"status": "ERROR", "error": msg}
        
        if not openpyxl:
            return {"status": "ERROR", "error": "Biblioteca openpyxl não instalada."}

        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
            data = {}
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                data[sheet] = [[cell.value for cell in row] for row in ws.iter_rows()]
            
            return {
                "status": "SUCCESS",
                "conteudo": str(data),
                "metadados": {"planilhas": wb.sheetnames, "formato": "XLSX"}
            }
        except Exception as e:
            return {"status": "ERROR", "error": str(e)}

def executar(entrada):
    return ExtratorXLSX().executar(entrada)
